from flask import Flask, request, jsonify, send_file
import pandas as pd
import io
import tempfile
from fpdf import FPDF
from openpyxl import Workbook

app = Flask(__name__)

@app.route("/parse", methods=["POST"])
def parse():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file provided"}), 400

    # Simule parsing Excel → JSON aligné prod
    data = {
        "prenom": "Charles",
        "modele": "Marie",
        "spc": {
            "score": 78,
            "details": {
                "conversion": 30,
                "effort": 24,
                "exploitable": 24
            },
            "pondération": [40, 30, 30]
        },
        "kpis": {
            "sales_net": 388.0,
            "minutes": 195,
            "fans_actifs": 21,
            "ppv_envoyés": 29,
            "ppv_achetés": 6,
            "unlocks": 6,
            "messages": 168,
            "ppv_ratio": 0.21,
            "unlock_ratio": 0.29,
            "prix_moyen": 3.95,
            "golden_ratio": 0.157,
            "ca_horaire": 119.4
        },
        "flags": {
            "actifs": ["prix moyen trop bas", "volume sans conversion"],
            "silencieux": ["peu de push", "CA horaire instable"]
        },
        "typologies": ["push sec", "volumeur inefficace"],
        "axe": "Travailler la perception de valeur via la montée progressive",
        "modules_recommandés": [3, 10, 13],
        "appel_managérial": {
            "recommandé": True,
            "motif": "Pricing incohérent malgré effort"
        },
        "salaire": {
            "brut": 58.2,
            "ajustement": 0,
            "net": 58.2
        },
        "date_semaine": "2025-07-01",
        "context_model": {
            "traffics_model": "modéré",
            "CA_objectif_chat": 400,
            "statut": "normal",
            "bonus_dispo": False
        }
    }

    return jsonify(data)

@app.route("/generate_pdf", methods=["POST"])
def generate_pdf():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON payload provided"}), 400

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt=f"Rapport Chatteur : {data['prenom']}", ln=1)
        pdf.cell(200, 10, txt=f"Modèle : {data['modele']}", ln=1)
        pdf.cell(200, 10, txt=f"SPC : {data['spc']['score']} /100", ln=1)
        pdf.cell(200, 10, txt=f"Axe : {data['axe']}", ln=1)
        pdf.cell(200, 10, txt=f"Modules à revoir : {', '.join(str(m) for m in data['modules_recommandés'])}", ln=1)
        pdf.cell(200, 10, txt=f"Salaire net : {data['salaire']['net']}€", ln=1)

        temp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        pdf.output(temp.name)
        return send_file(temp.name, as_attachment=True, download_name=f"rapport_{data['prenom']}.pdf")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/generate_xlsx", methods=["POST"])
def generate_xlsx():
    try:
        json_list = request.get_json()
        if not isinstance(json_list, list):
            return jsonify({"error": "Expected a list of JSON chatteurs"}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = "Synthèse Manager"

        headers = [
            "Chatteur", "Modèle", "SPC", "Sales NET", "$/h", "Typo 1", "Typo 2",
            "Flags", "Axe", "Modules", "Salaire brut", "Ajustement", "Salaire final", "Appel ?"
        ]
        ws.append(headers)

        for data in json_list:
            row = [
                data.get("prenom"),
                data.get("modele"),
                data["spc"]["score"],
                data["kpis"]["sales_net"],
                data["kpis"]["ca_horaire"],
                data["typologies"][0] if len(data["typologies"]) > 0 else "",
                data["typologies"][1] if len(data["typologies"]) > 1 else "",
                ", ".join(data["flags"]["actifs"]),
                data["axe"],
                ", ".join(str(m) for m in data["modules_recommandés"]),
                data["salaire"]["brut"],
                data["salaire"].get("ajustement", 0),
                data["salaire"]["net"],
                "Oui" if data["appel_managérial"]["recommandé"] else "Non"
            ]
            ws.append(row)

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        return send_file(temp_file.name, as_attachment=True, download_name="synthese_manager.xlsx")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
